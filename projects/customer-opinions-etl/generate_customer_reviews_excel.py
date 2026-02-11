"""
Generate a mock Excel file with customer reviews for a bank mobile app.
Creates realistic-looking customer feedback data for portfolio mockup purposes.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

# Output file
OUTPUT_FILE = "customer_reviews_q4_2024.xlsx"

# Review sources
SOURCES = [
    "App Store",
    "Google Play",
    "Customer Survey",
    "Social Media",
    "In-App Feedback",
]

# Reviewer first names (generic, fictional)
FIRST_NAMES = [
    "Anna", "Piotr", "Marta", "Tomasz", "Kasia", "Michał", "Ewa", "Paweł",
    "Agnieszka", "Krzysztof", "Monika", "Marcin", "Joanna", "Adam", "Magda",
    "Łukasz", "Natalia", "Jakub", "Aleksandra", "Bartek", "Justyna", "Rafał",
    "Karolina", "Dawid", "Patrycja", "Wojtek", "Sylwia", "Artur", "Dominika",
]

# Review templates by rating category
POSITIVE_REVIEWS = [
    "The app is really intuitive. I can do all my banking in seconds!",
    "Love the new design update. Much cleaner and easier to navigate.",
    "BLIK payments work flawlessly. Best feature ever!",
    "Fast login with fingerprint. Very secure and convenient.",
    "Finally a bank app that actually works well. Recommended!",
    "The transaction history is very detailed. I can track everything easily.",
    "Push notifications for transactions are super helpful.",
    "Transferring money to family has never been easier.",
    "The app rarely crashes. Very stable compared to other bank apps.",
    "Customer support chat in the app is very responsive.",
    "I love being able to block my card instantly from the app.",
    "Currency exchange rates are competitive and clearly displayed.",
    "The savings goals feature helped me save for vacation!",
    "Quick balance check on the home screen is perfect.",
    "Setting up recurring payments was surprisingly simple.",
]

NEUTRAL_REVIEWS = [
    "The app works fine but could use some improvements in speed.",
    "It's okay. Does what it needs to do for basic banking.",
    "Some features are good, others need work. Average experience.",
    "The app is functional but the design feels a bit dated.",
    "Works most of the time. Occasional slowdowns during peak hours.",
    "Basic features are there. Would like to see more analytics.",
    "It's a standard banking app. Nothing special but gets the job done.",
    "The login process could be faster. Otherwise acceptable.",
    "Not bad, but I've seen better mobile banking apps.",
    "Meets basic needs. Hoping for more features in future updates.",
]

NEGATIVE_REVIEWS = [
    "App keeps logging me out. Very frustrating experience.",
    "Too many steps to make a simple transfer. Needs simplification.",
    "The app crashed when I tried to pay a bill. Lost my progress.",
    "Loading times are terrible. Takes forever to check balance.",
    "The new update broke several features. Please fix ASAP!",
    "Can't find the option to export transaction history to PDF.",
    "Fingerprint login stopped working after the last update.",
    "The app drains my battery too quickly.",
    "Push notifications don't work reliably. Missed important alerts.",
    "The investment section is confusing and hard to understand.",
    "Why do I need to re-login every time I switch apps?",
    "Error messages are not helpful. Just says 'something went wrong'.",
]


def generate_reviews():
    """Generate mock customer reviews."""

    reviews = []
    start_date = datetime(2024, 10, 1)
    end_date = datetime(2024, 12, 31)
    total_days = (end_date - start_date).days

    review_id = 1001

    # Generate reviews for Q4 2024
    for _ in range(150):  # ~150 reviews over 3 months
        review_date = start_date + timedelta(days=random.randint(0, total_days))

        # Rating distribution: more positive than negative (realistic for decent app)
        rating_weights = [5, 10, 15, 35, 35]  # 1-star to 5-star weights
        rating = random.choices([1, 2, 3, 4, 5], weights=rating_weights)[0]

        # Select review text based on rating
        if rating >= 4:
            review_text = random.choice(POSITIVE_REVIEWS)
            sentiment = "Positive"
        elif rating == 3:
            review_text = random.choice(NEUTRAL_REVIEWS)
            sentiment = "Neutral"
        else:
            review_text = random.choice(NEGATIVE_REVIEWS)
            sentiment = "Negative"

        # Random source
        source = random.choice(SOURCES)

        # Reviewer name (some anonymous)
        if random.random() > 0.3:
            reviewer = f"{random.choice(FIRST_NAMES)} {chr(random.randint(65, 90))}."
        else:
            reviewer = "Anonymous"

        # NPS category based on rating (simplified mapping)
        if rating >= 4:
            nps_category = "Promoter"
        elif rating == 3:
            nps_category = "Passive"
        else:
            nps_category = "Detractor"

        reviews.append({
            "review_id": f"REV-{review_id}",
            "date": review_date.strftime("%Y-%m-%d"),
            "reviewer": reviewer,
            "source": source,
            "rating": rating,
            "review_text": review_text,
            "sentiment": sentiment,
            "nps_category": nps_category,
        })

        review_id += 1

    # Sort by date
    reviews.sort(key=lambda x: x["date"])

    return reviews


def create_excel():
    """Create the Excel file with customer reviews."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Reviews"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Conditional fills for sentiment/rating
    positive_fill = PatternFill(start_color="d4efdf", end_color="d4efdf", fill_type="solid")
    neutral_fill = PatternFill(start_color="fdebd0", end_color="fdebd0", fill_type="solid")
    negative_fill = PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type="solid")

    # Headers
    headers = [
        "Review ID",
        "Date",
        "Reviewer",
        "Source",
        "Rating",
        "Review Text",
        "Sentiment",
        "NPS Category"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Generate reviews
    reviews = generate_reviews()

    # Counters for summary
    sentiment_counts = {"Positive": 0, "Neutral": 0, "Negative": 0}
    rating_counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    source_counts = {s: 0 for s in SOURCES}

    row = 2
    for review in reviews:
        ws.cell(row=row, column=1, value=review["review_id"]).alignment = center_align
        ws.cell(row=row, column=2, value=review["date"]).alignment = center_align
        ws.cell(row=row, column=3, value=review["reviewer"]).alignment = center_align
        ws.cell(row=row, column=4, value=review["source"]).alignment = center_align

        # Rating with star representation
        rating_cell = ws.cell(row=row, column=5, value=review["rating"])
        rating_cell.alignment = center_align

        # Review text with wrapping
        text_cell = ws.cell(row=row, column=6, value=review["review_text"])
        text_cell.alignment = wrap_align

        # Sentiment with conditional coloring
        sentiment_cell = ws.cell(row=row, column=7, value=review["sentiment"])
        sentiment_cell.alignment = center_align
        if review["sentiment"] == "Positive":
            sentiment_cell.fill = positive_fill
        elif review["sentiment"] == "Neutral":
            sentiment_cell.fill = neutral_fill
        else:
            sentiment_cell.fill = negative_fill

        # NPS Category
        nps_cell = ws.cell(row=row, column=8, value=review["nps_category"])
        nps_cell.alignment = center_align

        # Apply borders
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border

        # Update counters
        sentiment_counts[review["sentiment"]] += 1
        rating_counts[review["rating"]] += 1
        source_counts[review["source"]] += 1

        row += 1

    # Adjust column widths
    column_widths = [12, 12, 18, 16, 8, 60, 12, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.sheet_properties.tabColor = "1a5276"

    # Summary styles
    summary_header_font = Font(bold=True, size=12)

    # Rating distribution
    ws_summary.cell(row=1, column=1, value="Rating Distribution").font = summary_header_font
    ws_summary.cell(row=2, column=1, value="Rating")
    ws_summary.cell(row=2, column=2, value="Count")
    ws_summary.cell(row=2, column=3, value="Percentage")

    total_reviews = len(reviews)
    for i, rating in enumerate([5, 4, 3, 2, 1], start=3):
        ws_summary.cell(row=i, column=1, value=f"{rating} Stars")
        ws_summary.cell(row=i, column=2, value=rating_counts[rating])
        ws_summary.cell(row=i, column=3, value=f"{rating_counts[rating]/total_reviews*100:.1f}%")

    # Sentiment distribution
    ws_summary.cell(row=9, column=1, value="Sentiment Distribution").font = summary_header_font
    ws_summary.cell(row=10, column=1, value="Sentiment")
    ws_summary.cell(row=10, column=2, value="Count")
    ws_summary.cell(row=10, column=3, value="Percentage")

    for i, sentiment in enumerate(["Positive", "Neutral", "Negative"], start=11):
        ws_summary.cell(row=i, column=1, value=sentiment)
        ws_summary.cell(row=i, column=2, value=sentiment_counts[sentiment])
        ws_summary.cell(row=i, column=3, value=f"{sentiment_counts[sentiment]/total_reviews*100:.1f}%")

    # Source distribution
    ws_summary.cell(row=16, column=1, value="Source Distribution").font = summary_header_font
    ws_summary.cell(row=17, column=1, value="Source")
    ws_summary.cell(row=17, column=2, value="Count")

    for i, source in enumerate(SOURCES, start=18):
        ws_summary.cell(row=i, column=1, value=source)
        ws_summary.cell(row=i, column=2, value=source_counts[source])

    # Adjust summary column widths
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 12

    # Average rating
    avg_rating = sum(r * c for r, c in rating_counts.items()) / total_reviews
    ws_summary.cell(row=24, column=1, value="Average Rating").font = summary_header_font
    ws_summary.cell(row=24, column=2, value=round(avg_rating, 2))

    # NPS calculation (simplified)
    promoters = sentiment_counts["Positive"]
    detractors = sentiment_counts["Negative"]
    nps_score = ((promoters - detractors) / total_reviews) * 100

    ws_summary.cell(row=26, column=1, value="NPS Score").font = summary_header_font
    ws_summary.cell(row=26, column=2, value=round(nps_score, 1))

    # Save
    wb.save(OUTPUT_FILE)
    print(f"Excel file generated: {OUTPUT_FILE}")
    print(f"Total reviews: {total_reviews}")
    print(f"Average rating: {avg_rating:.2f}")
    print(f"NPS Score: {nps_score:.1f}")
    print(f"\nSentiment breakdown:")
    for sentiment, count in sentiment_counts.items():
        print(f"  {sentiment}: {count} ({count/total_reviews*100:.1f}%)")


if __name__ == "__main__":
    create_excel()
