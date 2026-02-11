"""
Generate a mock Excel file representing a convoluted Credit Risk Entity Report.
This creates multiple small tables with formatting, pie charts, and bar charts,
plus additional empty sheets to show the complexity of the original process.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
import random

OUTPUT_FILE = "entity_report_q4_2023.xlsx"

# Color scheme for formatting
COLORS = {
    "header_dark": "1a365d",
    "header_mid": "2c5282",
    "header_light": "3182ce",
    "accent_green": "38a169",
    "accent_red": "e53e3e",
    "accent_yellow": "d69e2e",
    "light_gray": "edf2f7",
    "white": "ffffff",
}

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FONT = Font(bold=True, color="1a365d", size=9)
DATA_FONT = Font(size=9)
THIN_BORDER = Border(
    left=Side(style='thin', color='a0aec0'),
    right=Side(style='thin', color='a0aec0'),
    top=Side(style='thin', color='a0aec0'),
    bottom=Side(style='thin', color='a0aec0')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


def apply_header_style(cell, color_key="header_dark"):
    cell.font = HEADER_FONT
    cell.fill = PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type="solid")
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGN


def apply_data_style(cell, is_number=False, highlight=None):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = RIGHT_ALIGN if is_number else LEFT_ALIGN
    if highlight == "green":
        cell.fill = PatternFill(start_color="c6f6d5", end_color="c6f6d5", fill_type="solid")
    elif highlight == "red":
        cell.fill = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")
    elif highlight == "yellow":
        cell.fill = PatternFill(start_color="fefcbf", end_color="fefcbf", fill_type="solid")


def create_main_sheet(wb):
    """Create the main sheet with multiple convoluted tables and charts."""
    ws = wb.active
    ws.title = "Executive Summary"

    # ===== TABLE 1: Portfolio Overview (Top-left) =====
    ws.merge_cells('B2:E2')
    ws['B2'] = "Portfolio Overview"
    apply_header_style(ws['B2'])

    overview_headers = ["Metric", "Current", "Previous", "Change %"]
    overview_data = [
        ["Total Exposure (M PLN)", 45_892, 43_567, 5.3],
        ["NPL Ratio (%)", 2.8, 3.1, -9.7],
        ["Coverage Ratio (%)", 68.5, 65.2, 5.1],
        ["RWA (M PLN)", 28_456, 27_890, 2.0],
        ["Expected Loss (M PLN)", 892, 945, -5.6],
    ]

    for col, header in enumerate(overview_headers, 2):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell, "header_mid")

    for row_idx, row_data in enumerate(overview_data, 4):
        for col_idx, value in enumerate(row_data, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            is_num = col_idx > 2
            highlight = None
            if col_idx == 5 and isinstance(value, (int, float)):
                highlight = "green" if value > 0 else "red" if value < 0 else None
            apply_data_style(cell, is_number=is_num, highlight=highlight)
            if col_idx in [3, 4] and row_idx > 4:
                cell.number_format = '#,##0'
            elif col_idx == 5:
                cell.number_format = '0.0'

    # ===== TABLE 2: Risk Rating Distribution (Top-right) =====
    ws.merge_cells('G2:J2')
    ws['G2'] = "Risk Rating Distribution"
    apply_header_style(ws['G2'])

    rating_headers = ["Rating", "Count", "Exposure (M)", "% Total"]
    rating_data = [
        ["AAA-AA", 245, 8_920, 19.4],
        ["A", 512, 12_450, 27.1],
        ["BBB", 834, 15_670, 34.2],
        ["BB", 356, 6_234, 13.6],
        ["B & Below", 189, 2_618, 5.7],
    ]

    for col, header in enumerate(rating_headers, 7):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell, "header_mid")

    for row_idx, row_data in enumerate(rating_data, 4):
        for col_idx, value in enumerate(row_data, 7):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell, is_number=col_idx > 7)
            if col_idx in [8, 9]:
                cell.number_format = '#,##0'
            elif col_idx == 10:
                cell.number_format = '0.0'

    # ===== PIE CHART: Rating Distribution =====
    pie = PieChart()
    labels = Reference(ws, min_col=7, min_row=4, max_row=8)
    data = Reference(ws, min_col=9, min_row=3, max_row=8)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Exposure by Rating"
    pie.width = 10
    pie.height = 7
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    ws.add_chart(pie, "L2")

    # ===== TABLE 3: Segment Breakdown (Middle-left) =====
    ws.merge_cells('B11:F11')
    ws['B11'] = "Segment Breakdown"
    apply_header_style(ws['B11'])

    segment_headers = ["Segment", "Clients", "Exposure", "NPL", "Coverage"]
    segment_data = [
        ["Large Corporate", 156, 18_450, 1.2, 72.5],
        ["SME", 1_245, 12_890, 3.8, 65.2],
        ["Micro Enterprise", 3_567, 8_234, 4.5, 58.9],
        ["Specialized Lending", 89, 4_120, 2.1, 78.3],
        ["Project Finance", 34, 2_198, 0.8, 82.1],
    ]

    for col, header in enumerate(segment_headers, 2):
        cell = ws.cell(row=12, column=col, value=header)
        apply_header_style(cell, "header_mid")

    for row_idx, row_data in enumerate(segment_data, 13):
        for col_idx, value in enumerate(row_data, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell, is_number=col_idx > 2)
            if col_idx in [3, 4]:
                cell.number_format = '#,##0'
            elif col_idx in [5, 6]:
                cell.number_format = '0.0'

    # ===== BAR CHART: Segment Exposure =====
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Exposure by Segment"
    bar.y_axis.title = "M PLN"
    data = Reference(ws, min_col=4, min_row=12, max_row=17)
    cats = Reference(ws, min_col=2, min_row=13, max_row=17)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.shape = 4
    bar.width = 12
    bar.height = 7
    ws.add_chart(bar, "G11")

    # ===== TABLE 4: Key Risk Indicators (Middle-right) =====
    ws.merge_cells('B20:E20')
    ws['B20'] = "Key Risk Indicators"
    apply_header_style(ws['B20'])

    kri_headers = ["Indicator", "Threshold", "Actual", "Status"]
    kri_data = [
        ["NPL Ratio", "< 4.0%", "2.8%", "OK"],
        ["LTV > 100%", "< 5.0%", "3.2%", "OK"],
        ["Single Name Conc.", "< 3.0%", "2.1%", "OK"],
        ["Sector Concentration", "< 15.0%", "14.8%", "WARN"],
        ["FX Exposure", "< 25.0%", "18.5%", "OK"],
        ["Covenant Breaches", "< 2.0%", "2.3%", "BREACH"],
    ]

    for col, header in enumerate(kri_headers, 2):
        cell = ws.cell(row=21, column=col, value=header)
        apply_header_style(cell, "header_mid")

    for row_idx, row_data in enumerate(kri_data, 22):
        for col_idx, value in enumerate(row_data, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            highlight = None
            if col_idx == 5:
                if value == "OK":
                    highlight = "green"
                elif value == "WARN":
                    highlight = "yellow"
                elif value == "BREACH":
                    highlight = "red"
            apply_data_style(cell, is_number=False, highlight=highlight)

    # ===== TABLE 5: Monthly Trend (Bottom-left) =====
    ws.merge_cells('G20:L20')
    ws['G20'] = "Monthly Exposure Trend (M PLN)"
    apply_header_style(ws['G20'])

    months = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly_values = [42_890, 43_120, 43_567, 44_230, 45_120, 45_892]

    for col, month in enumerate(months, 7):
        cell = ws.cell(row=21, column=col, value=month)
        apply_header_style(cell, "header_light")

    for col, value in enumerate(monthly_values, 7):
        cell = ws.cell(row=22, column=col, value=value)
        apply_data_style(cell, is_number=True)
        cell.number_format = '#,##0'

    # ===== TABLE 6: Vintage Analysis (Bottom) =====
    ws.merge_cells('B30:H30')
    ws['B30'] = "Vintage Analysis - Default Rates by Origination Year"
    apply_header_style(ws['B30'])

    vintage_headers = ["Vintage", "2019", "2020", "2021", "2022", "2023", "Total Exp"]
    vintage_data = [
        ["Year 1", 0.5, 0.8, 0.4, 0.6, 0.3, 8_920],
        ["Year 2", 1.2, 1.5, 1.1, 0.9, "-", 12_450],
        ["Year 3", 2.1, 2.4, 1.8, "-", "-", 15_670],
        ["Year 4", 2.8, 2.9, "-", "-", "-", 6_234],
        ["Year 5+", 3.2, "-", "-", "-", "-", 2_618],
    ]

    for col, header in enumerate(vintage_headers, 2):
        cell = ws.cell(row=31, column=col, value=header)
        apply_header_style(cell, "header_mid")

    for row_idx, row_data in enumerate(vintage_data, 32):
        for col_idx, value in enumerate(row_data, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            is_num = col_idx > 2 and value != "-"
            apply_data_style(cell, is_number=is_num)
            if col_idx == 8 and isinstance(value, int):
                cell.number_format = '#,##0'
            elif is_num and col_idx < 8:
                cell.number_format = '0.0'

    # ===== TABLE 7: Watch List Summary (Bottom-right) =====
    ws.merge_cells('J30:N30')
    ws['J30'] = "Watch List Summary"
    apply_header_style(ws['J30'])

    watchlist_headers = ["Category", "Count", "Exposure", "% Port", "Action"]
    watchlist_data = [
        ["Early Warning", 45, 1_234, 2.7, "Monitor"],
        ["Under Review", 23, 890, 1.9, "Assess"],
        ["Restructured", 12, 456, 1.0, "Follow-up"],
        ["Recovery", 8, 234, 0.5, "Workout"],
    ]

    for col, header in enumerate(watchlist_headers, 10):
        cell = ws.cell(row=31, column=col, value=header)
        apply_header_style(cell, "header_mid")

    for row_idx, row_data in enumerate(watchlist_data, 32):
        for col_idx, value in enumerate(row_data, 10):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell, is_number=col_idx in [11, 12, 13])
            if col_idx == 12:
                cell.number_format = '#,##0'
            elif col_idx == 13:
                cell.number_format = '0.0'

    # Adjust column widths
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_additional_sheets(wb):
    """Create additional empty sheets to show complexity."""
    sheet_names = [
        "Large Corporate",
        "SME Portfolio",
        "Micro Enterprises",
        "Sector Analysis",
        "Geographic Distribution",
        "Collateral Coverage",
        "Provisions Movement",
        "IFRS9 Staging",
        "Concentration Risk",
        "FX Exposure",
        "Maturity Profile",
        "Data Sources",
        "SQL Queries",
        "Calculation Notes",
        "Archive Q3",
    ]

    for name in sheet_names:
        ws = wb.create_sheet(title=name)
        # Add a placeholder header to make it look like it has content
        ws['B2'] = f"{name} - Data"
        ws['B2'].font = Font(bold=True, size=12, color="718096")
        ws['B3'] = "Source: Multiple SQL queries and Excel files"
        ws['B3'].font = Font(italic=True, size=9, color="a0aec0")


def main():
    wb = openpyxl.Workbook()

    create_main_sheet(wb)
    create_additional_sheets(wb)

    wb.save(OUTPUT_FILE)
    print(f"Excel file generated: {OUTPUT_FILE}")
    print(f"Total sheets: {len(wb.sheetnames)}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
