"""
Generate a mockup Excel file for the Financial Data Reconciliation Tool project.
Creates an account balance reconciliation report with balances from multiple systems
and conditional formatting to highlight discrepancies.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule


def create_reconciliation_data():
    """Create sample account reconciliation data with intentional discrepancies."""

    accounts = [
        # Current Assets
        {"Account Code": "1100", "Account Name": "Cash and Cash Equivalents", "Category": "Current Assets"},
        {"Account Code": "1110", "Account Name": "Petty Cash", "Category": "Current Assets"},
        {"Account Code": "1120", "Account Name": "Checking Account - Operating", "Category": "Current Assets"},
        {"Account Code": "1200", "Account Name": "Accounts Receivable", "Category": "Current Assets"},
        {"Account Code": "1210", "Account Name": "Accounts Receivable - Trade", "Category": "Current Assets"},
        {"Account Code": "1220", "Account Name": "Accounts Receivable - Other", "Category": "Current Assets"},
        {"Account Code": "1250", "Account Name": "Allowance for Doubtful Accounts", "Category": "Current Assets"},
        {"Account Code": "1300", "Account Name": "Inventory", "Category": "Current Assets"},
        {"Account Code": "1310", "Account Name": "Raw Materials", "Category": "Current Assets"},
        {"Account Code": "1320", "Account Name": "Work in Progress", "Category": "Current Assets"},
        {"Account Code": "1330", "Account Name": "Finished Goods", "Category": "Current Assets"},
        {"Account Code": "1400", "Account Name": "Prepaid Expenses", "Category": "Current Assets"},
        {"Account Code": "1410", "Account Name": "Prepaid Insurance", "Category": "Current Assets"},
        {"Account Code": "1420", "Account Name": "Prepaid Rent", "Category": "Current Assets"},

        # Fixed Assets
        {"Account Code": "1500", "Account Name": "Property, Plant & Equipment", "Category": "Fixed Assets"},
        {"Account Code": "1510", "Account Name": "Land", "Category": "Fixed Assets"},
        {"Account Code": "1520", "Account Name": "Buildings", "Category": "Fixed Assets"},
        {"Account Code": "1530", "Account Name": "Machinery & Equipment", "Category": "Fixed Assets"},
        {"Account Code": "1540", "Account Name": "Vehicles", "Category": "Fixed Assets"},
        {"Account Code": "1550", "Account Name": "Furniture & Fixtures", "Category": "Fixed Assets"},
        {"Account Code": "1600", "Account Name": "Accumulated Depreciation", "Category": "Fixed Assets"},

        # Liabilities
        {"Account Code": "2100", "Account Name": "Accounts Payable", "Category": "Current Liabilities"},
        {"Account Code": "2110", "Account Name": "Accounts Payable - Trade", "Category": "Current Liabilities"},
        {"Account Code": "2120", "Account Name": "Accounts Payable - Other", "Category": "Current Liabilities"},
        {"Account Code": "2200", "Account Name": "Accrued Expenses", "Category": "Current Liabilities"},
        {"Account Code": "2210", "Account Name": "Accrued Salaries & Wages", "Category": "Current Liabilities"},
        {"Account Code": "2220", "Account Name": "Accrued Interest", "Category": "Current Liabilities"},
        {"Account Code": "2300", "Account Name": "Short-term Debt", "Category": "Current Liabilities"},
        {"Account Code": "2400", "Account Name": "Deferred Revenue", "Category": "Current Liabilities"},
        {"Account Code": "2500", "Account Name": "Long-term Debt", "Category": "Long-term Liabilities"},
        {"Account Code": "2510", "Account Name": "Bank Loans Payable", "Category": "Long-term Liabilities"},

        # Equity
        {"Account Code": "3100", "Account Name": "Common Stock", "Category": "Equity"},
        {"Account Code": "3200", "Account Name": "Additional Paid-in Capital", "Category": "Equity"},
        {"Account Code": "3300", "Account Name": "Retained Earnings", "Category": "Equity"},
        {"Account Code": "3400", "Account Name": "Treasury Stock", "Category": "Equity"},
    ]

    # Define balances - some matching, some with small differences, some with big differences
    balances = [
        # Current Assets - mix of matches and mismatches
        (2_450_000.00, 2_450_000.00),    # Match
        (5_000.00, 5_000.00),             # Match
        (2_445_000.00, 2_445_000.00),     # Match
        (1_875_340.50, 1_875_340.50),     # Match (subtotal)
        (1_650_000.00, 1_647_500.00),     # Small diff (2,500)
        (225_340.50, 227_840.50),         # Small diff (2,500)
        (-45_000.00, -45_000.00),         # Match
        (3_250_000.00, 3_315_000.00),     # Big diff (65,000)
        (875_000.00, 875_000.00),         # Match
        (425_000.00, 490_000.00),         # Big diff (65,000)
        (1_950_000.00, 1_950_000.00),     # Match
        (185_000.00, 185_000.00),         # Match
        (125_000.00, 125_000.00),         # Match
        (60_000.00, 60_000.00),           # Match

        # Fixed Assets
        (12_500_000.00, 12_500_000.00),   # Match (subtotal)
        (2_500_000.00, 2_500_000.00),     # Match
        (5_750_000.00, 5_750_000.00),     # Match
        (3_250_000.00, 3_250_000.00),     # Match
        (650_000.00, 655_000.00),         # Small diff (5,000)
        (350_000.00, 345_000.00),         # Small diff (5,000)
        (-4_125_000.00, -4_125_000.00),   # Match

        # Liabilities
        (1_450_000.00, 1_523_750.00),     # Big diff (73,750)
        (1_250_000.00, 1_250_000.00),     # Match
        (200_000.00, 273_750.00),         # Big diff (73,750)
        (425_000.00, 425_000.00),         # Match
        (275_000.00, 276_500.00),         # Small diff (1,500)
        (150_000.00, 148_500.00),         # Small diff (1,500)
        (500_000.00, 500_000.00),         # Match
        (175_000.00, 175_000.00),         # Match
        (3_500_000.00, 3_500_000.00),     # Match
        (3_500_000.00, 3_500_000.00),     # Match

        # Equity
        (1_000_000.00, 1_000_000.00),     # Match
        (2_500_000.00, 2_500_000.00),     # Match
        (8_750_340.50, 8_750_340.50),     # Match
        (-250_000.00, -250_000.00),       # Match
    ]

    # Build the data with system balances and differences
    data = []
    for i, account in enumerate(accounts):
        system_y, system_z = balances[i]
        diff = round(system_y - system_z, 2)

        data.append({
            "Account Code": account["Account Code"],
            "Account Name": account["Account Name"],
            "Category": account["Category"],
            "IBM Planning Analytics": system_y,
            "OneStream": system_z,
            "Difference": diff,
        })

    return pd.DataFrame(data)


def apply_formatting(ws):
    """Apply formatting to the worksheet."""

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    number_format = '#,##0.00'

    # Conditional formatting colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15

    # Format data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Number formatting for columns D, E, F (system balances and difference)
            if cell.column >= 4:
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Add conditional formatting for the Difference column (column F)
    # Green for exact match (0)
    ws.conditional_formatting.add(
        'F2:F100',
        FormulaRule(
            formula=['$F2=0'],
            fill=green_fill
        )
    )

    # Yellow for small differences (absolute value <= 10000)
    ws.conditional_formatting.add(
        'F2:F100',
        FormulaRule(
            formula=['AND($F2<>0, ABS($F2)<=10000)'],
            fill=yellow_fill
        )
    )

    # Red for big differences (absolute value > 10000)
    ws.conditional_formatting.add(
        'F2:F100',
        FormulaRule(
            formula=['ABS($F2)>10000'],
            fill=red_fill
        )
    )

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Set row height for header
    ws.row_dimensions[1].height = 30


def create_summary_sheet(wb, df):
    """Create a summary sheet with reconciliation statistics."""

    ws = wb.create_sheet("Summary")

    # Calculate statistics
    total_accounts = len(df)
    matched = len(df[df['Difference'] == 0])
    small_diff = len(df[(df['Difference'] != 0) & (abs(df['Difference']) <= 10000)])
    big_diff = len(df[abs(df['Difference']) > 10000])

    total_diff = df['Difference'].sum()

    # Build summary data
    summary_data = [
        ["RECONCILIATION SUMMARY", ""],
        ["", ""],
        ["Report Date:", "January 2020"],
        ["Systems Compared:", "IBM Planning Analytics vs OneStream"],
        ["", ""],
        ["STATISTICS", ""],
        ["Total Accounts Reviewed:", total_accounts],
        ["Accounts Matched (Green):", matched],
        ["Small Discrepancies (Yellow):", small_diff],
        ["Large Discrepancies (Red):", big_diff],
        ["", ""],
        ["Total Difference:", total_diff],
        ["", ""],
        ["DISCREPANCY THRESHOLDS", ""],
        ["Match (Green):", "Difference = $0"],
        ["Small (Yellow):", "Difference <= $10,000"],
        ["Large (Red):", "Difference > $10,000"],
    ]

    for row in summary_data:
        ws.append(row)

    # Apply formatting
    header_font = Font(bold=True, size=14, color="2F5496")
    section_font = Font(bold=True, size=11)

    ws['A1'].font = header_font
    ws['A6'].font = section_font
    ws['A14'].font = section_font

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35

    # Format the total difference with number format
    ws['B12'].number_format = '#,##0.00'


def main():
    """Generate the reconciliation Excel file."""

    # Create data
    df = create_reconciliation_data()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Reconciliation"

    # Write data to worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Apply formatting
    apply_formatting(ws)

    # Create summary sheet
    create_summary_sheet(wb, df)

    # Save the file
    output_path = "/Users/przemyslawkepka/Desktop/GIT_NEW/pk-data-sol-website-mockups/financial-data-reconciliation/account_reconciliation_jan2020.xlsx"
    wb.save(output_path)
    print(f"Excel file created: {output_path}")

    # Print summary
    print("\nReconciliation Summary:")
    print(f"Total accounts: {len(df)}")
    print(f"Matched: {len(df[df['Difference'] == 0])}")
    print(f"Small discrepancies: {len(df[(df['Difference'] != 0) & (abs(df['Difference']) <= 10000)])}")
    print(f"Large discrepancies: {len(df[abs(df['Difference']) > 10000])}")


if __name__ == "__main__":
    main()
